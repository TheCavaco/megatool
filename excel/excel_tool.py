import openpyxl
import os
import glob



import xlrd
from xlwt import Workbook
from xlutils.copy import copy

from database.excel_db import ExcelDB




def read_invoices_in_chronological_order(directory:str) -> list:
    # List all subdirectories in the given directory
    subdirs = [d for d in os.listdir(directory) if os.path.isdir(os.path.join(directory, d))]

    # Filter only those directories that are named after years (assuming names are valid years)
    year_dirs = [d for d in subdirs if d.isdigit() and len(d) == 4]

    
    # Sort the directories by year (which are strings representing the year)
    year_dirs.sort()
    
    chronological_order_excel: list = []

    for year_dir in year_dirs:
        

        year_dir_path = os.path.join(directory, year_dir)




        dir_listed:list = os.listdir(year_dir_path)
        # Get a sorted copy of the list
        sorted_dir_listed = sorted(dir_listed)

        # Now print the sorted list

        

        for dir in sorted_dir_listed:
            search_pattern = os.path.join(directory, year_dir, dir,  '*INVOICE*.xls')
            matching_files = glob.glob(search_pattern, recursive=False)
            # If there is at least one matching file
            if matching_files:
                chronological_order_excel += [matching_files[0]]
            pass
    return chronological_order_excel



class ExcelTool():
    def __init__(self):
        self.workbook = None
        pass

    def load_workbook(self, filename:str, read: bool, old:bool):
        if old:
            # Use xlrd
            self.workbook = xlrd.open_workbook(filename)
            pass
        else:
            # Use openpyxl
            self.workbook = openpyxl.load_workbook(filename, read_only=read)

    def read_values(self, directory:str, database:ExcelDB, old:bool=True):
        """ 
        Read all values from all the available items in the data folder
        """
        excel_lst: list = read_invoices_in_chronological_order(directory)

        # marks : 17, 0....
        # references: 17,11....
        for file in excel_lst:
            self.load_workbook(file, True, old)
            if old:
                sheet = self.workbook.sheet_by_index(0)
                
                # Starting index for Hugo's files
                index:int = 17
                try:
                    mark = sheet.cell_value(index, 0)
                    reference= sheet.cell_value(index, 11)
                except IndexError:
                    print("To be done file: " + file)
                    continue

                while mark != "":
                    if reference != "X":
                        # References to be manually altered are skipped
                        try:
                            if isinstance(mark, float):
                                database.add_entry(str(int(mark)), int(reference))
                            else:
                                database.add_entry(str(mark), int(reference))
                        except Exception as e:
                            print("Found X reference or mark")
                            print(e)

                    index += 1
                    try:
                        mark = sheet.cell_value(index, 0)
                        reference= sheet.cell_value(index, 11)
                    except IndexError:
                        print("Finished file: " + file)
                        break
            else:
                # TODO: not implemented
                sheet = self.workbook['INVOICE']
                value = sheet['A17'].value
                print(f"Value in A17: {value}")
                return value
        print("Added all to table.")
    

    def populate(self, filename:str, database:ExcelDB, old:bool=True):
        if old:
            # Reading the file
            sheet = self.workbook.sheet_by_index(0)

            # Copy the content to a new writable workbook
            workbook_wr = copy(self.workbook)

            # Select the sheet you want to edit
            sheet_wr = workbook_wr.get_sheet(0)  # 0 is the index of the first sheet


            index = 17
            current_item = 1

            reference_dict: dict = {}


            passed_by_x:True = False

            mark: str = ""

            try:
                # Read mark from file
                mark = sheet.cell_value(index, 0)
            except IndexError:
                print("Error in file: " + filename)
                return
            
            while mark != "":
                reference:int = 0
                # Search in Database
                if isinstance(mark, float):
                    reference = database.get_entry(str(int(mark)))
                else:
                    reference = database.get_entry(str(mark))

                if reference == "X":
                    passed_by_x = True

                # Write in file
                sheet_wr.write(index, 11, reference)
                
                if not passed_by_x:
                    if reference in reference_dict.keys():
                        sheet_wr.write(index, 10, reference_dict[reference])
                    else:
                        reference_dict[reference] = current_item
                        current_item += 1


                # New mark
                index += 1

                try:
                    # Read mark from file
                    temp_mark = sheet.cell_value(index, 0)

                    if temp_mark != mark:
                        current_item +=1

                    mark = temp_mark
                except IndexError:
                    print("File completed successfully")
                    break

                pass

            workbook_wr.save('test_workbook.xls')
            print("File completed successfully")




            pass
        else:
            #not implemented
            pass

    pass